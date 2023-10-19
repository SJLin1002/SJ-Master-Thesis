# 依時間
import random
import time
import copy
import openpyxl
from openpyxl.styles import Alignment



def create_car(car_num): 
    data = []
    for i in range(car_num):
        # data增加car_count台車,[x,y,v]copy
        data.append(
            [random.randint(0, 100), random.choice((0, 3)), random.randint(15, 30)]
            #車輛時速54~108km/hr(15~30m/s)
        )
    return data

#更新data,1秒?! 
def actual_change_car_data(dataA):
    for car in dataA:
        #車輛X座標改變
        car[0] = round((car[0] + car[2]),3)  
        k = random.random()
        #更改Y(車道) 80%改變車道
        if k > 0.2 :
            if car[1] == 0:
                car[1] = 3
            else: car[1] = 0      
        #改變車輛速度  (10% ~ -10%)
        a = random.uniform(0.1,-0.1) 
        New_v = car[2] * (1+a)
        car[2] = round(New_v,3) 
    # print(data)
    return dataA

def simulation_change_car_data(dataS):
    for car in dataS:
        #車輛X座標改變
        car[0] = round((car[0] + car[2]),3)        
    return dataS

# 兩點中點(新)
def kmeans(data):
    #kmeans 
    # clusters=2，隨機生成兩點
    node1=[round(random.uniform(0,100),3),round(random.uniform(0,3),3)]
    node2=[round(random.uniform(0,100),3),round(random.uniform(0,3),3)]
    while True:
        clusA=[]
        clusB=[]    
        #車(data)到node1,node2的距離-分群
        for i in range(len(data)):
            d1=((data[i][0]-node1[0])**2+(data[i][1]-node1[1])**2)**0.5
            d2=((data[i][0]-node2[0])**2+(data[i][1]-node2[1])**2)**0.5
            if d1 < d2 :
                clusA.append([data[i][0],data[i][1]])
            else: 
                clusB.append([data[i][0],data[i][1]])
            # print("d1 :",d1)
            # print("d2 :",d2)
        #重心點
        if len(clusA)==0 :
            node1=[round(random.uniform(0,100),3),round(random.uniform(0,3),3)]
            #print("A群=None")
            continue
        elif len(clusB)==0:
            node2=[round(random.uniform(0,100),3),round(random.uniform(0,3),3)]
            #print("B群=None")
            continue
        else:
            Ax = 0
            Ay = 0
            Bx = 0
            By = 0

            for j in range(len(clusA)):
                Ax += clusA[j][0]
                Ay += clusA[j][1]
            for k in range(len(clusB)):
                Bx += clusB[k][0]
                By += clusB[k][1]

            New_node1 = (round(Ax/len(clusA),3),round(Ay/len(clusA),3))
            New_node2 = (round(Bx/len(clusB),3),round(By/len(clusB),3))
         
            if node1 == New_node1 and node2 == New_node2:
                # print("A :",clusA,"Node1 :",New_node1)
                # print("B :",clusB,"Node2 :",New_node2)
                break
            else:
                node1=New_node1
                node2=New_node2
    # print("A :",clusA,"Node1 :",New_node1)
    # print("B :",clusB,"Node2 :",New_node2) 
    close = 1000
    for q in range(len(clusA)):
        for w in range(len(clusB)):
            dis = ((clusA[q][0]-clusB[w][0])**2+(clusA[q][1]-clusB[w][1])**2)**0.5
            if dis < close :
                close = dis
                closeA = clusA[q]
                closeB = clusB[w]
    # center_point = ((node1[0]+node2[0])/2,(node1[1]+node2[1])/2)  
    # print(closeA,closeB)
    center_point = ((closeA[0]+closeB[0])/2, (closeA[1]+closeB[1])/2)
    return center_point          

def act(data,location_UAV):
    time_step = 0
    times  = 1
    #經delaytime，車子的位置及車輛中心點
    initial_center= kmeans(data)
    d=((location_UAV[0]-initial_center[0])**2+(location_UAV[1]-initial_center[1])**2)**0.5
    sheet['B2'] = 0 
    sheet['C2'] = d
    print("初始中心點 :", initial_center)
   
    for i in range(1, int(delaytTime/0.5)+1):  
        # 更新車輛位置
        data = actual_change_car_data(data)
        # 在每段時間，找到車輛中心點
        center_point = kmeans(data)
        # print(f"CenterPoint at Time {i}:", center_point)
       
    
    print()
    print("Time 0 :",delaytTime,"s" )
    # print("Updated Car Positions at Time 0:", data)
    print("車輛中心點 :", center_point)
    print("無人機位置 :",location_UAV)
    print()


    while True:
        sheet['A'+ str(times+2)] = times
        print(f"Time {times}:", time_step + delaytTime,"秒")
        
        # 更新車輛位置
        data = actual_change_car_data(data)
        # print(f"Updated Car Positions at Time {time_step}:", data)

        # 車輛位置的中心點
        center_point = kmeans(data)
       
        print(f"CenterPoint at Time {times}:", center_point)

        # 無人機位置
        location_UAV=((location_UAV[0]+location_UAV[2]),location_UAV[1],location_UAV[2])
        print("無人機位置 : ",location_UAV)
        # 無人機與車輛中心點距離
        d=((location_UAV[0]-center_point[0])**2+(location_UAV[1]-center_point[1])**2)**0.5
        print("A距離 :",d)

        #在Excel 紀錄秒數和距離
        sheet['B'+ str(times+2)] = time_step + delaytTime
        sheet['C'+ str(times+2)] = d
        # if d<5:
        if location_UAV[0] > center_point[0]:
            print("act",time_step + delaytTime,"秒 OKOK") 
            break
        time_step+=0.5
        times += 1
        # time.sleep(0.5)
        
        print()
    
    return time_step + delaytTime

# 模擬時間循環
def sim(data,location_UAV):
    time_step = 0
    times = 1

    #經delaytime，車子的位置及車輛中心點
    initial_center= kmeans(data)
    d=((location_UAV[0]-initial_center[0])**2+(location_UAV[1]-initial_center[1])**2)**0.5
    sheet['D2'] = 0 
    sheet['E2'] = d
    print("初始中心點 :", initial_center)
   
    
    for i in range(1, int(delaytTime/0.5)+1): 
        # 更新車輛位置
        data = simulation_change_car_data(data)
        # 在每段時間，找到車輛中心點
        center_point = kmeans(data)
      
        # print(f"CenterPoint at Time {i}:", center_point)
    
    
   
    print()
    print("Time 0:",delaytTime,"s" )
    # print("Updated Car Positions at Time 0:", data)
    print("CenterPoint:", center_point)
    print("無人機位置 :",location_UAV)
    print()

    while True:
        sheet['A'+ str(times+2)] = times
        print(f"Time {times}:", time_step + delaytTime,"秒")
        
        # 更新車輛位置
        data =simulation_change_car_data(data)
        # print(f"Updated Car Positions at Time {time_step}:", data)
        
        # 車輛位置的中心點
        center_point = kmeans(data)
        
        print(f"CenterPoint at Time {times}:", center_point)


        location_UAV=((location_UAV[0]+location_UAV[2]),location_UAV[1],location_UAV[2])
        print("無人機位置 : ",location_UAV)
        # 無人機與車輛中心點距離
        d=((location_UAV[0]-center_point[0])**2+(location_UAV[1]-center_point[1])**2)**0.5
        print("S距離 :",d)

        #在Excel 紀錄秒數和距離
        sheet['D'+ str(times+2)] = time_step + delaytTime
        sheet['E'+ str(times+2)] = d

        # if d<5:
        if location_UAV[0] > center_point[0]:
            print("sim",time_step + delaytTime,"秒 OKOK") 
            break
        time_step += 0.5
        times+=1
        # time.sleep(0.5)
        print()
   
    return time_step + delaytTime

def nopred(data,location_UAV):
    time_step = 0.5
    times  = 1

    initial_center= kmeans(data)
    d=((location_UAV[0]-initial_center[0])**2+(location_UAV[1]-initial_center[1])**2)**0.5
    sheet['F2'] = 0 
    sheet['G2'] = d
    print("初始中心點 :", initial_center)

    print()
    print("Time 0:",time_step,"s" )
    # print("Updated Car Positions at Time 0:", data)
    print("CenterPoint:", initial_center)
    print("無人機位置 :",location_UAV)
    print()

    while True:
        sheet['A'+ str(times+2)] = times
        print(f"Time {times}:", time_step,"秒")
        
        # 更新車輛位置
        data =simulation_change_car_data(data)
        # print(f"Updated Car Positions at Time {time_step}:", data)
        
        # 車輛位置的中心點
        center_point = kmeans(data)
        print(f"CenterPoint at Time {times}:", center_point)


        location_UAV=((location_UAV[0]+location_UAV[2]),location_UAV[1],location_UAV[2])
        print("無人機位置 : ",location_UAV) 
        # 無人機與車輛中心點距離
        d=((location_UAV[0]-center_point[0])**2+(location_UAV[1]-center_point[1])**2)**0.5
        print("N距離 :",d)

        #在Excel 紀錄秒數和距離
        sheet['F'+ str(times+2)] = time_step 
        sheet['G'+ str(times+2)] = d

        # if d<5:
        if location_UAV[0] > center_point[0]:
            print("nopred ",time_step," OKOK") 
            break
        time_step+=0.5
        times+=1
        # time.sleep(0.5)
        print()
    return time_step


#使用openpyxl 內 Workbook 方法建立一個新的工作簿
workbook = openpyxl.Workbook()
#取得第一個工作表
sheet = workbook.worksheets[0]
sheet['A1'] = "Time"
sheet['A2'] = 0
sheet.merge_cells('B1:C1')
sheet.merge_cells('D1:E1')
sheet.merge_cells('F1:G1')
sheet['B1'] = "Act"
sheet['D1'] = "Sim"
sheet['F1'] = "Nopred"


car_num = 10
# O_data=create_car(car_num)
O_data = [[36, 0, 19], [88, 3, 22], [53, 3, 20], [65, 3, 16], [18, 3, 18], [24, 3, 21], [92, 3, 27], [37, 3, 23], [91, 0, 26], [50, 3, 24]]


dataS = copy.deepcopy(O_data)
dataA = copy.deepcopy(O_data)   
dataN = copy.deepcopy(O_data)

print("初始車輛位置 : ",O_data) 
print()
#無人機時速126km/hr(35m/s)
V_UAV = 35
delaytTime = 3 #(s)
location_UAV=(0,0,V_UAV)  #無人機初始位置

act = act(dataA,location_UAV)
print()
sim = sim(dataS,location_UAV)
print()
# nopred = nopred(dataN,location_UAV)

print("act: ",act,"秒")
print("sim: ",sim,"秒")
# print("nopred",nopred,"秒")

for i in range(1,sheet.max_row):
    for j in range(1,sheet.max_column):
        sheet.cell(row=i+1,column=j+1).alignment=Alignment(vertical='center',   horizontal='center') 
        # sheet[str([i-1][j-1])].alignment=Alignment(vertical='center', horizontal='left') 
workbook.save('data_Compare .xlsx')