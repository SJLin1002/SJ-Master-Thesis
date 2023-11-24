
import random
import time
import copy
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

def all(delaytTime,car_num,road):
    def create_car(car_num): 
        data = []
        for i in range(car_num):
            # data增加car_count台車,[x,y,v]copy
            if road == 0 :
                data.append(
                    #平面道路-車輛時速10~60km/hr(2.777~16.66m/s)
                    [round(random.uniform(0, 100),3), random.choice((0, 3)), round(random.uniform(2.777,16.66),3)])
            else :    
                data.append(
                    #高速公路-車輛時速80~110km/hr(22.222~30.55m/s)
                    [round(random.uniform(0, 100),3), random.choice((0, 3)), round(random.uniform(22.222,30.55),3)])
            
            
        return data

    #更新data,每timeslot秒?! 
    def actual_change_car_data(dataA):
        for car in dataA:
            #車輛X座標改變
            car[0] = round((car[0] + (car[2]*timeslot)),3)  
            k = random.random()
            #更改Y(車道) 80%改變車道
            if k > 0.2 :
                if car[1] == 0:
                    car[1] = 3
                else: car[1] = 0      
            #改變車輛速度  (10% ~ -10%)
            a = random.uniform(0.1,-0.1) 
            New_v = car[2] * (1+a)
            car[2] = round(New_v,3) 
        # print(data)
        return dataA

    def simulation_change_car_data(dataS):
        for car in dataS:
            #車輛X座標改變
            car[0] = round((car[0] + (car[2]*timeslot)),3)        
        return dataS
    
    '''
    def constant_speed(data,time):
        #車輛在時間time後的位置(等速)
        for car in data:
            car[0] = round(car[0] + (car[2] * time))
        return data
    '''

    # 兩點中點(新)
    def kmeans(data):
            max = 0
            min = 100000
            for car in data:            
                if car[0] > max : max = car[0] 
                if car[0] < min : min = car[0]
            # clusters=2，隨機生成兩點
            node1=[round(random.uniform(min,max),3),round(random.uniform(0,3),3)]
            node2=[round(random.uniform(min,max),3),round(random.uniform(0,3),3)]
            while True:
                clusA=[]
                clusB=[]    
                #車(data)到node1,node2的距離-分群
                for i in range(len(data)):
                    d1=((data[i][0]-node1[0])**2+(data[i][1]-node1[1])**2)**0.5
                    d2=((data[i][0]-node2[0])**2+(data[i][1]-node2[1])**2)**0.5
                    if d1 < d2 :
                        clusA.append([data[i][0],data[i][1]])
                    else: 
                        clusB.append([data[i][0],data[i][1]])
                    # print("d1 :",d1)
                    # print("d2 :",d2)
                #重心點
                if len(clusA)==0 :
                    node1=[round(random.uniform(min,max),3),round(random.uniform(0,3),3)]
                    #print("A群=None")
                    continue
                elif len(clusB)==0:
                    node2=[round(random.uniform(min,max),3),round(random.uniform(0,3),3)]
                    #print("B群=None")
                    continue
                else:
                    Ax = 0
                    Ay = 0
                    Bx = 0
                    By = 0

                    for j in range(len(clusA)):
                        Ax += clusA[j][0]
                        Ay += clusA[j][1]
                    for k in range(len(clusB)):
                        Bx += clusB[k][0]
                        By += clusB[k][1]

                    New_node1 = (round(Ax/len(clusA),3),round(Ay/len(clusA),3))
                    New_node2 = (round(Bx/len(clusB),3),round(By/len(clusB),3))
                
                    if node1 == New_node1 and node2 == New_node2:
                        # print("A :",clusA,"Node1 :",New_node1)
                        # print("B :",clusB,"Node2 :",New_node2)
                        break
                    else:
                        node1=New_node1
                        node2=New_node2
            # print("A :",clusA,"Node1 :",New_node1)
            # print("B :",clusB,"Node2 :",New_node2) 
            close = 100000
            for q in range(len(clusA)):
                for w in range(len(clusB)):
                    dis = ((clusA[q][0]-clusB[w][0])**2+(clusA[q][1]-clusB[w][1])**2)**0.5
                    if dis < close :
                        close = dis
                        closeA = clusA[q]
                        closeB = clusB[w]
            # center_point = ((node1[0]+node2[0])/2,(node1[1]+node2[1])/2)  
            # print(closeA,closeB)
            center_point = ((closeA[0]+closeB[0])/2, (closeA[1]+closeB[1])/2)
            return center_point
    
    '''
    def act(dataA):
        # 更新車輛位置(每timeslot秒)
        dataA = actual_change_car_data(dataA)
        # 在每段時間，找到車輛中心點
        center_point = kmeans(dataA)
        return center_point
    '''
    # 車輛找到中心後，更新位置
    def sim(dataS):
        center_point = kmeans(dataS)
        dataS = simulation_change_car_data(dataS)
        return center_point
    # 無人機位置更新
    def uav(location_UAV):
        location_UAV = (location_UAV[0]+location_UAV[2]*timeslot,location_UAV[1],location_UAV[2])
        return location_UAV

    def SimAct(location_UAV,dataA):
        Num = 0
        location_UAV0 = location_UAV
        center_point = kmeans(dataA)
        print("center_point:",center_point)
        du = ((center_point[0]**2-location_UAV0[0]**2)+(center_point[0]**2-location_UAV0[0]**2))**0.5
        t = du/V_UAV

        while True : 
            #無人機delaytTime秒後追
            print("Num :",Num)
            if Num > delaytTime:
                location_UAV = uav(location_UAV)
            print("UAV : ",location_UAV)

            # 車輛每timeslot秒跑一次
            s = sim(dataS)
            print("sim : ",s)

            #d2 無人機追上sim時
            if s[0] < location_UAV[0] :
                center_pointS  = s #記下無人機追上dataS中點時，center_point的位置
                tS = Num #記下無人機追上dataS中點的時間
                print("SIM okok")
                break
    
            print()
            Num += timeslot 
        print("D1初始的位置 : ",center_point)
        print("D1初始UAV到center_point的時間 : ",t)

        print("D2追到的位置 : ",center_pointS)
        print("D2追到的時間 : ",tS)
        
        #計算ACT
        if int(tS/timeslot)>=int((t+delaytTime)/timeslot):
            for i in range(int(tS/timeslot)):  # D2
                dataA = actual_change_car_data(dataA)           
                if i == (int((t+delaytTime)/timeslot)-1) :# D1
                    New_center = kmeans(dataA)
            center_pointA = kmeans(dataA) 
        else:
            for i in range(int((t+delaytTime)/timeslot)):  # D1
                dataA = actual_change_car_data(dataA)           
                if i == (int(tS/timeslot)-1) :# D2
                    center_pointA = kmeans(dataA)
            New_center = kmeans(dataA) 

        print(t+delaytTime,"秒時，實際車群中心位置 :",New_center)
        print(tS,"秒時，實際車群中心位置 :",center_pointA)

        #D1 車輛群經過delaytimec(3秒)後，實際車輛群的位置(a-latency_S)
        d1 = ((New_center[0]-center_point[0])**2+(New_center[1]-center_point[1])**2)**0.5

        #D2 無人機追上時SIM時，ACT實際位置(s - center_pointS )
        
        d2 =((center_pointA[0]-center_pointS [0])**2+(center_pointA[1]-center_pointS [1])**2)**0.5

        print("d1距離 : ",d1)
        print("d1時間 : ",t+delaytTime)
        print("d2距離 : ",d2)
        print("d2時間 : ",tS)

        return d1,d2


    
    O_data = create_car(car_num)
    # O_data = [[36, 0, 19], [88, 3, 22], [53, 3, 20], [65, 3, 16], [18, 3, 18], [24, 3, 21], [92, 3, 27], [37, 3, 23], [91, 0, 26], [50, 3, 24]]
    timeslot = 0.5 #時間間隔
    dataA = copy.deepcopy(O_data)
    dataS = copy.deepcopy(O_data)  

    
    V_UAV = 44.32
    # UAV最高速160km/hr(44.32m/s)
    # delaytTime = 3 #(s)
    location_UAV=(0,0,V_UAV)  #無人機初始位置
    

    d1,d2 = SimAct(location_UAV,dataA)

    print("D1 = ",d1,"D2 = ",d2)
    print()
    return d1,d2




# #使用openpyxl 內 Workbook 方法建立一個新的工作簿
# workbook = openpyxl.Workbook()
# 開啟Excel文件
workbook = openpyxl.load_workbook('template.xlsx')
#取得第一個工作表
sheet = workbook.worksheets[0]
n = 50
times = 0
D1, D2 = [], []
car_num = 25
road = 0  # road = 0 平面道路，road = 1 高速公路
for delayTime in range(0, 11):
    D1_total = 0
    D2_total = 0
    
    sheet.cell(column=1 + times*4, row=1).value = f'Delay Time = {delayTime}'
    sheet.cell(column=2 + times*4, row=1).value = datetime.now().strftime("%H:%M:%S")

    sheet.cell(column=1 + times*4, row=2).value = 'No.'
    sheet.cell(column=2 + times*4, row=2).value = 'D1無預測'
    sheet.cell(column=3 + times*4, row=2).value = 'D2模擬'
    
    for i in range(1,n+1):
        k = all(delayTime,car_num,road)
        sheet.cell(column=1 + times*4, row=i + 2).value = i
        sheet.cell(column=2 + times*4, row=i + 2).value = k[0]
        sheet.cell(column=3 + times*4, row=i + 2).value = k[1]

        D1_total += k[0]
        D2_total += k[1]

    D1.append(D1_total/n)
    D2.append(D2_total/n)

    times += 1
# print(D1)
# print(D2)
# sheet = workbook.create_sheet("Mysheet")
sheet = workbook.worksheets[1]

sheet.cell(column=1, row=1).value = 'DelayTime'
sheet.cell(column=2, row=1).value = 'D1無預測'
sheet.cell(column=3, row=1).value = 'D2模擬'

for i in range(0, 11):
    sheet.cell(column=1, row=i+2).value = i

times = 2
for i in D1:
    sheet.cell(column=2, row=times).value = round(i,2) 
    times += 1

times = 2
for i in D2:
    sheet.cell(column=3, row=times).value = round(i,2) 
    times += 1
g = "General road"
h = "Highway"
# workbook.save(f'data/{datetime.now().strftime("%H%M%S")}.xlsx')

# workbook.save(f'data/{h,car_num}.xlsx')
workbook.save(f'data/{h}{car_num}.xlsx')

# road = 0 平面道路(G)，road = 1 高速公路(H)
if road == 0 : 
    workbook.save(f'data/{g}{car_num}.xlsx')
    print(g)
else : 
    workbook.save(f'data/{h}{car_num}.xlsx')
    print(h)

print("CarNum : ",car_num)
print("n : ",n)